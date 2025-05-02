from openpyxl import load_workbook
import pandas as pd
from typing import List, Dict, Optional, Tuple
import os
from datetime import datetime, timedelta
import sqlite3
from config import CONFIG

ruta_excel = r"C:\Users\alias\OneDrive\Escritorio\AIEP\Semestre 6\Automatizacion\BOT-FACTURA\data\facturas.xlsx"
print(f"Existe archivo: {os.path.exists(ruta_excel)}")
print(f"Es legible: {os.access(ruta_excel, os.R_OK)}")

def procesar_factura_excel(ruta: str) -> Dict:
    try:
        wb = load_workbook(ruta, data_only=True)
        sheet = wb.active
        
        fecha_celda = sheet['C2'].value
        fecha_emision = None  # Valor por defecto

        if fecha_celda:
            try:
                if hasattr(fecha_celda, 'strftime'):  # Si es objeto fecha
                    fecha_obj = fecha_celda
                else:  # Si es string
                    # Intenta parsear desde formato DD-MM-YYYY
                    fecha_obj = datetime.strptime(str(fecha_celda), "%d-%m-%Y")
                
                # Convertir a formato SQLite (YYYY-MM-DD)
                fecha_emision = fecha_obj.strftime("%Y-%m-%d")

                 # Validación de fecha no futura y no mayor a 2 años
                hoy = datetime.now().date()
                fecha_factura = fecha_obj.date()
                max_antiguedad = hoy - timedelta(days=730)  # 2 años
                if fecha_factura > hoy:
                    raise ValueError("Fecha de factura no puede ser futura")
                if fecha_factura < max_antiguedad:
                    raise ValueError("Fecha de factura no puede tener más de 2 años de antigüedad")
                
            except (ValueError, AttributeError) as e:
                print(f"⚠️ Error al procesar fecha (celda C2): {str(e)}")
       

        # 1. Encabezado (ajusta las celdas a tu Excel real)
        encabezado = {
            "numero_factura": sheet['A2'].value,
            "proveedor_id": int(sheet['B2'].value) if sheet['B2'].value else None,
            "fecha_emision": fecha_emision,
            "comentarios": str(sheet['D2'].value) if sheet['D2'].value else None
        }
        
        # 2. Items (con while y control explícito)
        items = []
        row = 6
        max_row = sheet.max_row
        productos_vistos = set()  # Para validar items duplicados
        
        while row <= max_row:
            celda_producto = sheet[f'A{row}']
            celda_cantidad = sheet[f'B{row}']
            celda_precio = sheet[f'C{row}']
            
            # Condición de salida (si encuentra celda vacía o texto "total")
            if not celda_producto.value or "total" in str(celda_producto.value).lower():
                break
                
            # Procesar ítem solo si hay producto
            if celda_producto.value:
                try:
                    producto = str(celda_producto.value).strip()
                    cantidad = float(celda_cantidad.value) if celda_cantidad.value else 0.0
                    precio = float(celda_precio.value) if celda_precio.value else 0.0
                    
                    # Validación de cantidad (positiva y < 10,000)
                    if cantidad <= 0:
                        raise ValueError(f"Fila {row}: Cantidad debe ser positiva")
                    if cantidad >= 10000:
                        raise ValueError(f"Fila {row}: Cantidad excede límite (10,000)")
                    
                    # Validación de item duplicado
                    item_key = (producto.lower(), precio)
                    if item_key in productos_vistos:
                        raise ValueError(f"Fila {row}: Item duplicado ({producto} - ${precio})")
                    productos_vistos.add(item_key)
                    


                    items.append({
                        "producto": str(celda_producto.value).strip(),
                        "cantidad": float(celda_cantidad.value) if celda_cantidad.value else 0.0,
                        "precio_unitario": float(celda_precio.value) if celda_precio.value else 0.0
                    })
                except (ValueError, TypeError) as e:
                    print(f"Fila {row}: Error en datos - {e}")
            
            row += 1  # Incremento dentro del while!
        
        # Validación de número de factura duplicado
        if encabezado["numero_factura"]:
            try:
                conn = sqlite3.connect(CONFIG['ruta_bd'])
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM facturas WHERE numero_factura = ?", 
                             (encabezado["numero_factura"],))
                if cursor.fetchone() is not None:
                    raise ValueError(f"Factura {encabezado['numero_factura']} ya existe en el sistema")
            except sqlite3.Error as e:
                print(f"⚠️ Advertencia: No se pudo verificar duplicado en BD: {str(e)}")
            finally:
                conn.close()

        # 3. Totales (busca desde el final)
        totales = {"iva": 0.0, "monto_total": 0.0}
        # Obtener IVA de D13
        try:
            iva_celda = sheet['D13'].value
            if iva_celda is not None:
                totales["iva"] = float(iva_celda)
        except (ValueError, TypeError) as e:
            print(f"Error al convertir IVA: {e}")

        # Obtener Monto Total de D14
        try:
            total_celda = sheet['D14'].value
            if total_celda is not None:
                totales["monto_total"] = float(total_celda)
        except (ValueError, TypeError) as e:
            print(f"Error al convertir Monto Total: {e}")

        # Si no se encontraron valores, calcularlos como respaldo
        if totales["monto_total"] == 0.0 and items:
            subtotal = sum(item["cantidad"] * item["precio_unitario"] for item in items)
            totales["monto_total"] = subtotal + totales["iva"]

        return {
            "encabezado": encabezado,
            "items": items,
            "totales": totales
        }
        
    except Exception as e:
        raise ValueError(f"Error al procesar Excel: {str(e)}")

def validar_factura(datos_factura: Dict) -> List[str]:
    errores = []
    encabezado = datos_factura["encabezado"]
    items = datos_factura["items"]

    if not encabezado.get("numero_factura"):
        errores.append("Número de factura vacío")
    
    if not items:
        errores.append("La factura no contiene items")
    
    for i, item in enumerate(items, 1):
        if not item.get("producto"):
            errores.append(f"Item {i}: Producto no especificado")
        if not isinstance(item.get("cantidad"), (int, float)) or item["cantidad"] <= 0:
            errores.append(f"Item {i}: Cantidad inválida")
        if not isinstance(item.get("precio_unitario"), (int, float)) or item["precio_unitario"] <= 0:
            errores.append(f"Item {i}: Precio unitario inválido")
    
    return errores

def clasificar_factura(
    datos_factura: Dict, 
    historico_proveedor: Optional[Dict]
) -> Tuple[str, Optional[str]]:
    

    errores = validar_factura(datos_factura)
    if errores:
        return "Rechazada", "; ".join(errores)
    
    if not historico_proveedor:
        return "Pendiente", "Proveedor no registrado"
    
    monto_actual = datos_factura["totales"]["monto_total"]
    monto_historico = historico_proveedor.get("monto_promedio", 0)
    
    if abs(monto_actual - monto_historico) > monto_historico * 0.2:
        return "Pendiente", "Monto difiere del histórico en más del 20%"
    
    return "Aprobada", None